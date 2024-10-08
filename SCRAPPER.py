import re 
import asyncio
import os
from typing import Dict, Optional, List
import aiohttp
from telethon import TelegramClient, events
from telethon.tl.functions.channels import JoinChannelRequest
from telethon.errors import ChannelPrivateError
import logging
from openpyxl import load_workbook, Workbook
from datetime import datetime
import json
import http.client
from aiohttp import ClientTimeout
from telethon import Button
import io
import aiohttp
import emoji
from datetime import datetime
import sqlite3
from typing import List, Optional
from dataclasses import dataclass, field
from typing import Dict, Optional, List, Any

# Configuración de logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Configuración de la aplicación
API_ID = '28602320'  # Reemplaza con tus credenciales
API_HASH = 'f1a4624ba2dca49bfc6a30f7febf5ce6'
SOURCE_CHANNEL = 'AisakaScrapper'
DESTINATION_CHANNEL = 'https://t.me/Seya_scrapper'
SEYA_CHK_URL = "https://t.me/SeyaChk_bot"
IMAGE_URL = 'https://i.ibb.co/zXNVx1k/bi7jjmbd-1-1-1-1.png'

# Configuración de APIs
RAPIDAPI_HOST = "bin-ip-checker.p.rapidapi.com"
RAPIDAPI_KEY = "fd634539damsh43c186a88067f87p180bbdjsna0b20f8d5c1a"
BINLIST_API_URL = "https://lookup.binlist.net/"
BINCODES_API_URL = "https://api.bincodes.com/bin/"
BINCODES_API_KEY = "ab982b2b6d0a3632b69d97b09f2f4ee3"

# Archivos de base de datos local
BIN_DATABASE_FILE = r"C:\Checker\bin_database.xlsx"
BIN_CACHE_FILE = r"C:\Checker\bin_cache.json"

@dataclass

@dataclass
class CardInfo:
    cc: str
    mm: str
    yy: str
    cvv: str
    extras: List[str] = field(default_factory=list)
    card_type: str = ""
    bank: str = ""
    country: str = ""
    date: str = ""
    bin_info: Optional[Dict[str, Any]] = None


class BinDatabase:
    def __init__(self):
        self.excel_db = self.load_excel_database()
        self.cache = self.load_cache()

    def load_excel_database(self) -> Dict[str, Dict]:
        bin_database = {}
        if os.path.exists(BIN_DATABASE_FILE):
            try:
                workbook = load_workbook(BIN_DATABASE_FILE)
                sheet = workbook.active
                headers = [cell.value for cell in sheet[1]]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    row_data = dict(zip(headers, row))
                    if 'bin' in row_data and row_data['bin']:
                        bin_database[str(row_data['bin'])] = {
                            'scheme': row_data.get('scheme', ''),
                            'type': row_data.get('type', ''),
                            'brand': row_data.get('brand', ''),
                            'country': {'name': row_data.get('country', '')},
                            'bank': {'name': row_data.get('bank', '')},
                            'extras': json.loads(row_data.get('extras', '[]'))
                        }
                logger.info(f"Base de datos Excel cargada. Total de registros: {len(bin_database)}")
            except Exception as e:
                logger.error(f"Error al cargar la base de datos Excel: {str(e)}")
        else:
            logger.info(f"Archivo {BIN_DATABASE_FILE} no encontrado. Creando base de datos vacía.")
            self.create_bin_database()
        return bin_database

    def create_bin_database(self):
        try:
            workbook = Workbook()
            sheet = workbook.active
            headers = ['bin', 'scheme', 'type', 'brand', 'country', 'bank', 'extras']
            sheet.append(headers)
            workbook.save(BIN_DATABASE_FILE)
            logger.info(f"Archivo de base de datos BIN creado: {BIN_DATABASE_FILE}")
        except Exception as e:
            logger.error(f"Error al crear la base de datos BIN: {str(e)}")

    def load_cache(self) -> Dict[str, Dict]:
        if os.path.exists(BIN_CACHE_FILE):
            try:
                with open(BIN_CACHE_FILE, 'r') as f:
                    return json.load(f)
            except Exception as e:
                logger.error(f"Error al cargar el caché: {str(e)}")
        return {}

    def save_cache(self):
        try:
            with open(BIN_CACHE_FILE, 'w') as f:
                json.dump(self.cache, f)
        except Exception as e:
            logger.error(f"Error al guardar el caché: {str(e)}")

    def get_bin_info(self, bin_number: str) -> Optional[Dict]:
        if bin_number in self.cache:
            logger.info(f"BIN {bin_number} encontrado en caché")
            return self.cache[bin_number]
        
        if bin_number in self.excel_db:
            logger.info(f"BIN {bin_number} encontrado en base de datos Excel")
            bin_info = self.excel_db[bin_number]
            self.cache[bin_number] = bin_info
            self.save_cache()
            return bin_info
        
        logger.info(f"BIN {bin_number} no encontrado en la base de datos local")
        return None

    def save_bin_info(self, bin_number: str, bin_info: Dict):
        if not isinstance(bin_info, dict) or not any(bin_info.values()):
            logger.error(f"Información de BIN inválida para {bin_number}. Valor recibido: {bin_info}")
            return

        try:
            workbook = load_workbook(BIN_DATABASE_FILE) if os.path.exists(BIN_DATABASE_FILE) else Workbook()
            sheet = workbook.active
            
            if sheet.max_row == 0 or (sheet.max_row == 1 and not any(sheet[1])):
                headers = ['bin', 'scheme', 'type', 'brand', 'country', 'bank', 'extras']
                sheet.append(headers)
            
            row_data = [
                bin_number,
                bin_info.get('scheme', 'N/A'),
                bin_info.get('type', 'N/A'),
                bin_info.get('brand', 'N/A'),
                bin_info.get('country', {}).get('name', 'N/A'),
                bin_info.get('bank', {}).get('name', 'N/A'),
                json.dumps(bin_info.get('extras', []))
            ]
            
            bin_exists = False
            for row in sheet.iter_rows(min_row=2):
                if row[0].value == bin_number:
                    for i, value in enumerate(row_data):
                        row[i].value = value
                    bin_exists = True
                    break
            
            if not bin_exists:
                sheet.append(row_data)
            
            workbook.save(BIN_DATABASE_FILE)
            
            self.excel_db[bin_number] = bin_info
            self.cache[bin_number] = bin_info
            self.save_cache()
            
            logger.info(f"BIN {bin_number} guardado exitosamente en la base de datos y caché.")
        except Exception as e:
            logger.error(f"Error al guardar BIN en la base de datos: {str(e)}")
            logger.exception(e)

    def save_extras(self, bin_number: str, extras: List[str]):
        bin_info = self.get_bin_info(bin_number)
        if bin_info:
            bin_info['extras'] = extras
            self.save_bin_info(bin_number, bin_info)
        else:
            logger.warning(f"No se encontró información para el BIN {bin_number} al guardar extras")


async def query_binlist_api(bin_number: str) -> Optional[Dict]:
    try:
        async with aiohttp.ClientSession(timeout=ClientTimeout(total=10)) as session:
            async with session.get(f"{BINLIST_API_URL}{bin_number}") as response:
                if response.status == 200:
                    data = await response.json()
                    return {
                        'scheme': data.get('scheme', ''),
                        'type': data.get('type', ''),
                        'brand': data.get('brand', ''),
                        'country': {'name': data.get('country', {}).get('name', '')},
                        'bank': {'name': data.get('bank', {}).get('name', '')}
                    }
    except Exception as e:
        logger.error(f"Error de Binlist API para BIN {bin_number}: {str(e)}")
    return None

async def query_rapidapi_bin_checker(bin_number: str) -> Optional[Dict]:
    try:
        conn = http.client.HTTPSConnection(RAPIDAPI_HOST)
        payload = json.dumps({"bin": bin_number, "ip": "8.8.8.8"})
        headers = {
            'x-rapidapi-key': RAPIDAPI_KEY,
            'x-rapidapi-host': RAPIDAPI_HOST,
            'Content-Type': "application/json"
        }
        conn.request("POST", f"/?bin={bin_number}&ip=8.8.8.8", payload, headers)
        res = conn.getresponse()
        data = json.loads(res.read().decode("utf-8"))

        if data.get("success"):
            return {
                'scheme': data.get('BIN', {}).get('scheme', '').strip(),
                'type': data.get('BIN', {}).get('type', '').strip(),
                'brand': data.get('BIN', {}).get('brand', '').strip(),
                'country': {'name': data.get('BIN', {}).get('country', {}).get('name', '').strip()},
                'bank': {'name': data.get('BIN', {}).get('issuer', {}).get('name', '').strip()}
            }
    except Exception as e:
        logger.error(f"Error de RapidAPI para BIN {bin_number}: {str(e)}")
    return None

def is_card_info_valid(card_info: Dict) -> bool:
    """
    Verifica si la información de la tarjeta tiene suficientes datos válidos para ser enviada.
    """
    if not isinstance(card_info, dict):
        return False
    
    # Contadores para datos válidos
    valid_fields = 0
    
    # Verificar campos principales
    for field in ['scheme', 'type', 'brand']:
        if card_info.get(field) and card_info[field].strip():
            valid_fields += 1
    
    # Verificar país
    if (card_info.get('country') and 
        isinstance(card_info['country'], dict) and 
        card_info['country'].get('name') and 
        card_info['country']['name'].strip()):
        valid_fields += 1
    
    # Verificar banco
    if (card_info.get('bank') and 
        isinstance(card_info['bank'], dict) and 
        card_info['bank'].get('name') and 
        card_info['bank']['name'].strip()):
        valid_fields += 1
    
    # Requerir al menos 3 campos válidos para considerar la información como válida
    return valid_fields >= 3

import sqlite3
from typing import List

class ExtrasDatabase:
    def __init__(self, db_file: str = 'extras_database.sqlite'):
        self.db_file = db_file
        self.conn = sqlite3.connect(db_file)
        self.create_table()

    def create_table(self):
        cursor = self.conn.cursor()
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS extras (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bin TEXT,
            extra TEXT
        )
        ''')
        self.conn.commit()

    def insert_extras(self, bin_number: str, extras: List[str]):
        cursor = self.conn.cursor()
        for extra in extras:
            cursor.execute('INSERT INTO extras (bin, extra) VALUES (?, ?)', (bin_number, extra))
        self.conn.commit()

    def get_extras(self, bin_number: str) -> List[str]:
        cursor = self.conn.cursor()
        cursor.execute('SELECT extra FROM extras WHERE bin = ?', (bin_number,))
        return [row[0] for row in cursor.fetchall()]

    def close(self):
        self.conn.close()

def escape_markdown_v2_telegram(text: str) -> str:
    """Escapa caracteres especiales para Markdown V2 en Telegram."""
    special_chars = ['_', '*', '[', ']', '(', ')', '~', '`', '>', '#', '+', '-', '=', '|', '{', '}', '.', '!']
    for char in special_chars:
        text = text.replace(char, f'\\{char}')
    return text

def escape_html(text: str) -> str:
    """Escapa caracteres especiales para HTML en Telegram."""
    return text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

from datetime import datetime

# Updated country code mapping with all countries and territories
COUNTRY_CODE_MAP = {
    'United States': 'US',
    'United States of America': 'US',
    'Spain': 'ES',
    'Mexico': 'MX',
    'Argentina': 'AR',
    'Brazil': 'BR',
    'Canada': 'CA',
    'France': 'FR',
    'Germany': 'DE',
    'Japan': 'JP',
    'United Kingdom': 'GB',
    'Italy': 'IT',
    'China': 'CN',
    'India': 'IN',
    'Australia': 'AU',
    'Russia': 'RU',
    'South Korea': 'KR',
    'Thailand': 'TH',
    'Vietnam': 'VN',
    'Malaysia': 'MY',
    'Singapore': 'SG',
    'Indonesia': 'ID',
    'Philippines': 'PH',
    'Turkey': 'TR',
    'Switzerland': 'CH',
    'Sweden': 'SE',
    'Norway': 'NO',
    'Denmark': 'DK',
    'Finland': 'FI',
    'Netherlands': 'NL',
    'Belgium': 'BE',
    'Poland': 'PL',
    'Austria': 'AT',
    'New Zealand': 'NZ',
    'Ireland': 'IE',
    'Portugal': 'PT',
    'Greece': 'GR',
    'United Arab Emirates': 'AE',
    'Bahrain': 'BH',
    'Saudi Arabia': 'SA',
    'Qatar': 'QA',
    'Egypt': 'EG',
    'Chile': 'CL',
    'Colombia': 'CO',
    'Peru': 'PE',
    'South Africa': 'ZA',
    'Kenya': 'KE',
    'Nigeria': 'NG',
    'Morocco': 'MA',
    'Israel': 'IL',
    'Jordan': 'JO',
    'Lebanon': 'LB',
    'Kuwait': 'KW',
    'Oman': 'OM',
    'Pakistan': 'PK',
    'Bangladesh': 'BD',
    'Sri Lanka': 'LK',
    'Nepal': 'NP',
    'Afghanistan': 'AF',
    'Iran': 'IR',
    'Iraq': 'IQ',
    'Syria': 'SY',
    'Libya': 'LY',
    'Tunisia': 'TN',
    'Algeria': 'DZ',
    'Panama': 'PA',
    'Costa Rica': 'CR',
    'Guatemala': 'GT',
    'El Salvador': 'SV',
    'Honduras': 'HN',
    'Nicaragua': 'NI',
    'Venezuela': 'VE',
    'Paraguay': 'PY',
    'Uruguay': 'UY',
    'Ecuador': 'EC',
    'Bolivia': 'BO',
    'Cuba': 'CU',
    'Dominican Republic': 'DO',
    'Haiti': 'HT',
    'Jamaica': 'JM',
    'Trinidad and Tobago': 'TT',
    'Barbados': 'BB',
    'Bahamas': 'BS',
    'Iceland': 'IS',
    'Czech Republic': 'CZ',
    'Hungary': 'HU',
    'Romania': 'RO',
    'Bulgaria': 'BG',
    'Slovakia': 'SK',
    'Croatia': 'HR',
    'Slovenia': 'SI',
    'Serbia': 'RS',
    'Bosnia and Herzegovina': 'BA',
    'Montenegro': 'ME',
    'North Macedonia': 'MK',
    'Albania': 'AL',
    'Kosovo': 'XK',
    'Ukraine': 'UA',
    'Belarus': 'BY',
    'Moldova': 'MD',
    'Lithuania': 'LT',
    'Latvia': 'LV',
    'Estonia': 'EE',
    'Georgia': 'GE',
    'Armenia': 'AM',
    'Azerbaijan': 'AZ',
    'Kazakhstan': 'KZ',
    'Uzbekistan': 'UZ',
    'Turkmenistan': 'TM',
    'Kyrgyzstan': 'KG',
    'Tajikistan': 'TJ',
    'Malta': 'MT',
    'Cyprus': 'CY',
    'Luxembourg': 'LU',
    'Monaco': 'MC',
    'Liechtenstein': 'LI',
    'San Marino': 'SM',
    'Vatican City': 'VA',
    'Andorra': 'AD',
    'Greenland': 'GL',
    'Maldives': 'MV',
    'Fiji': 'FJ',
    'Papua New Guinea': 'PG',
    'Solomon Islands': 'SB',
    'Vanuatu': 'VU',
    'Samoa': 'WS',
    'Tonga': 'TO',
    'Micronesia': 'FM',
    'Marshall Islands': 'MH',
    'Palau': 'PW',
    'Zimbabwe': 'ZW',
    'Zambia': 'ZM',
    'Yemen': 'YE',
    'Western Sahara': 'EH',
    'Wallis and Futuna': 'WF',
    'Virgin Islands, U.S.': 'VI',
    'Virgin Islands, British': 'VG',
    'Uganda': 'UG',
    'Tuvalu': 'TV',
    'Turks and Caicos Islands': 'TC',
    'Taiwan': 'TW',
    'Sudan': 'SD',
    'South Sudan': 'SS',
    'Somalia': 'SO',
    'Sierra Leone': 'SL',
    'Seychelles': 'SC',
    'Senegal': 'SN',
    'São Tomé and Príncipe': 'ST',
    'Rwanda': 'RW',
    'Réunion': 'RE',
    'Puerto Rico': 'PR',
    'New Caledonia': 'NC',
    'Namibia': 'NA',
    'Mozambique': 'MZ',
    'Mayotte': 'YT',
    'Mauritius': 'MU',
    'Mauritania': 'MR',
    'Martinique': 'MQ',
    'Mali': 'ML',
    'Malawi': 'MW',
    'Madagascar': 'MG',
    'Macau': 'MO',
    'Liberia': 'LR',
    'Lesotho': 'LS',
    'Guinea-Bissau': 'GW',
    'Guinea': 'GN',
    'Guadeloupe': 'GP',
    'Ghana': 'GH',
    'Gambia': 'GM',
    'Gabon': 'GA',
    'French Polynesia': 'PF',
    'French Guiana': 'GF',
    'Ethiopia': 'ET',
    'Eritrea': 'ER',
    'Equatorial Guinea': 'GQ',
    'Djibouti': 'DJ',
    'Côte d\'Ivoire': 'CI',
    'Congo': 'CG',
    'Congo, Democratic Republic of the': 'CD',
    'Comoros': 'KM',
    'Central African Republic': 'CF',
    'Cape Verde': 'CV',
    'Cameroon': 'CM',
    'Burundi': 'BI',
    'Burkina Faso': 'BF',
    'British Indian Ocean Territory': 'IO',
    'Botswana': 'BW',
    'Benin': 'BJ',
    'Belize': 'BZ',
    'Bhutan': 'BT',
    'Bermuda': 'BM',
    'Anguilla': 'AI',
    'Antarctica': 'AQ',
    'American Samoa': 'AS'
}
def get_alpha2_code(country_name: str) -> str:
    """
    Get the ISO 3166-1 alpha-2 country code from country name.
    """
    if not country_name or not isinstance(country_name, str):
        return 'XX'
    
    # Normalize the country name
    normalized_name = country_name.strip().upper()
    
    # Direct lookup
    for key, code in COUNTRY_CODE_MAP.items():
        if key.upper() == normalized_name:
            return code
    
    # Partial match lookup
    for key, code in COUNTRY_CODE_MAP.items():
        if key.upper() in normalized_name or normalized_name in key.upper():
            return code
    
    # If no match is found, return XX
    return 'XX'



def get_country_flag(country_code: str) -> str:
    """
    Convert country code to flag emoji.
    """
    if not country_code or not isinstance(country_code, str):
        return "🏳️"
    
    country_code = country_code.strip().upper()
    if len(country_code) != 2 or not country_code.isalpha():
        return "🏳️"
    
    # Convert country code to regional indicator symbols
    try:
        flag = ''.join(chr(ord(c) + 127397) for c in country_code)
        return flag
    except Exception:
        return "🏳️"

from typing import Optional, Dict, Any
import re
from dataclasses import dataclass

def format_card_message(card: 'CardInfo') -> str:
    if not card.bin_info or not isinstance(card.bin_info, dict):
        card.bin_info = {
            'scheme': 'Desconocido',
            'type': 'Desconocido',
            'brand': 'Desconocido',
            'country': {'name': 'Desconocido'},
            'bank': {'name': 'Desconocido'}
        }
    
    bin_number = card.cc[:6]
    bank = card.bin_info.get('bank', {})
    bank_name = bank.get('name', 'Desconocido') if isinstance(bank, dict) else str(bank)
    country = card.bin_info.get('country', {})
    country_name = country.get('name', 'Desconocido') if isinstance(country, dict) else str(country)
    
    country_code = get_alpha2_code(country_name)
    country_flag = get_country_flag(country_code)
    
    scheme = card.bin_info.get('scheme', 'Desconocido').upper()
    card_type = card.bin_info.get('type', 'Desconocido').upper()
    brand = card.bin_info.get('brand', 'Desconocido').upper()
    
    # Mejora en el formato de extras
    extras_text = ""
    if card.extras:
        extras_text = "🔢 𝗘𝗫𝗧𝗥𝗔𝗦 🔢\n"
        extras_text += "━━━━━━━━━━━━━━━━━━━\n"
        for i, extra in enumerate(card.extras, 1):
            extras_text += f"└─❯ Extra{i}: <code>{extra}</code>\n"
    
    formatted_text = f"""
<b>╔══『 𝐒𝐞𝐲𝐚 𝐒𝐜𝐫𝐚𝐩𝐩𝐞𝐫 』══╗</b>
✅ TARJETA ENCONTRADA✅
━━━━━━━━━━━━━━━━━━━
└─❯ 💳 Tarjeta: <code>{card.cc}|{card.mm}|{card.yy}|{card.cvv}</code>
└─❯ 🎯 Bin: <code>{bin_number}</code>
└─❯ 🏦 Esquema: <code>{scheme}</code>
└─❯ ⚙️ Tipo: <code>{card_type}</code>
└─❯ ⭐️ Marca: <code>{brand}</code>
{extras_text}
└─❯ 🏛️ Banco: <code>{bank_name}</code>
└─❯ 🌎 País: <code>{country_name} {country_flag}</code>
⏰ 𝗧𝗜𝗘𝗠𝗣𝗢 𝗗𝗘 𝗖𝗛𝗘𝗤𝗨𝗘𝗢 ⏰
━━━━━━━━━━━━━━━━━━━
└─❯ ⏱️ Verificado: {card.date}
╚══════════════════════╝"""
    
    return formatted_text

def extract_card_info(text: str) -> Optional[CardInfo]:
    card_pattern = re.compile(r"Card » (\d{13,16})\|(\d{1,2})\|(\d{2,4})\|(\d{3,4})")
    extra_pattern = re.compile(r"Extra[¹²³]? » (.+)")  # Modificado para capturar cualquier contenido después de "Extra »"
    info_pattern = re.compile(r"Info » (.+)")
    bank_pattern = re.compile(r"Bank » (.+)")
    country_pattern = re.compile(r"Country » (.+)")
    date_pattern = re.compile(r"Date » (.+)")
    
    card_match = card_pattern.search(text)
    extra_matches = extra_pattern.findall(text)
    info_match = info_pattern.search(text)
    bank_match = bank_pattern.search(text)
    country_match = country_pattern.search(text)
    date_match = date_pattern.search(text)
    
    if card_match:
        cc, mm, yy, cvv = card_match.groups()
        extras = extra_matches  # Ahora guardamos todas las extras encontradas
        card_info = CardInfo(cc, mm, yy, cvv, extras)
        
        if info_match:
            card_info.card_type = info_match.group(1)
        if bank_match:
            card_info.bank = bank_match.group(1)
        if country_match:
            card_info.country = country_match.group(1)
        if date_match:
            card_info.date = date_match.group(1)
        
        return card_info
    return None
async def download_image(session: aiohttp.ClientSession) -> Optional[io.BytesIO]:
    """
    Descarga la imagen del banner desde la URL configurada.
    """
    try:
        async with session.get(IMAGE_URL) as response:
            if response.status == 200:
                image_content = await response.read()
                image = io.BytesIO(image_content)
                image.name = 'image.png'
                return image
            logger.error(f"No se pudo descargar la imagen. Código de estado: {response.status}")
            return None
    except Exception as e:
        logger.error(f"Error al descargar la imagen: {str(e)}")
        return None

async def send_card_message(client: TelegramClient, 
                          channel: str, 
                          message: str, 
                          image: io.BytesIO, 
                          button: List[List[Button]]):  # Cambiado para aceptar matriz de botones
    """
    Envía el mensaje formateado con la imagen y el botón al canal.
    """
    try:
        await client.send_file(
            entity=channel,
            file=image,
            caption=message,
            parse_mode='html',
            buttons=button,  # Ya es una matriz de botones
            force_document=False
        )
        logger.info(f"Mensaje enviado a {channel}")
    except Exception as e:
        logger.error(f"Error al enviar el mensaje: {str(e)}")
        logger.exception(e)  # Agregado para ver el error completo


async def process_card_info(client: TelegramClient, 
                            card_info: CardInfo, 
                            bin_database: BinDatabase, 
                            extras_db: ExtrasDatabase):
    """
    Procesa la información de la tarjeta, obtiene datos del BIN, y envía el mensaje formateado.
    """
    try:
        bin_number = card_info.cc[:6]
        logger.info(f"Procesando información para BIN: {bin_number}")
        
        # Obtener información del BIN desde la base de datos local
        card_info.bin_info = bin_database.get_bin_info(bin_number)
        
        # Si no hay información en la base de datos local, consultar APIs externas
        if not card_info.bin_info:
            logger.info(f"BIN {bin_number} no encontrado en base local, consultando APIs...")
            # Intentar primero con Binlist API
            card_info.bin_info = await query_binlist_api(bin_number)
            
            # Si Binlist falla, intentar con RapidAPI
            if not card_info.bin_info:
                card_info.bin_info = await query_rapidapi_bin_checker(bin_number)
            
            # Si se obtuvo información válida, guardarla en la base de datos
            if card_info.bin_info and is_card_info_valid(card_info.bin_info):
                logger.info(f"Guardando información de BIN {bin_number} en base de datos")
                bin_database.save_bin_info(bin_number, card_info.bin_info)

        # Verificar si tenemos información válida del BIN
        if not card_info.bin_info or not is_card_info_valid(card_info.bin_info):
            logger.warning(f"Información insuficiente para BIN {bin_number}. Mensaje no enviado.")
            return

        # Procesar y guardar extras
        if card_info.extras:
            logger.info(f"Guardando {len(card_info.extras)} extras para BIN {bin_number}")
            extras_db.insert_extras(bin_number, card_info.extras)

        # Obtener todas las extras históricas para este BIN
        all_extras = extras_db.get_extras(bin_number)
        card_info.extras = all_extras  # Actualizar con todas las extras disponibles

        # Formatear el mensaje
        card_message = format_card_message(card_info)
        if not card_message:
            logger.warning(f"No se pudo formatear el mensaje para BIN {bin_number}")
            return

        # Con esta nueva versión:

        # Crear los botones como una matriz
        buttons = [
            [
                Button.url("📊 𝐎𝐟𝐢𝐜𝐢𝐚𝐥 𝐂𝐡𝐚𝐧𝐧𝐞𝐥", "https://t.me/Seya_scrapper"),
                Button.url("🤖 𝐎𝐟𝐢𝐜𝐢𝐚𝐥 𝐁𝐨𝐭", SEYA_CHK_URL)
            ]
        ]


       # Enviar el mensaje con la imagen y los botones
        async with aiohttp.ClientSession() as session:
            image = await download_image(session)
            if image:
                try:
                    full_channel = DESTINATION_CHANNEL
                    if not full_channel.startswith('https://t.me/'):
                        full_channel = f'https://t.me/{full_channel.lstrip("@")}'
                    
                    await send_card_message(
                        client=client,
                        channel=full_channel,
                        message=card_message,
                        image=image,
                        button=buttons  # Aquí se usan los botones definidos arriba
                    )
                except ValueError as ve:
                    logger.error(f"Error de valor al enviar mensaje: {str(ve)}")
                except Exception as e:
                    logger.error(f"Error al enviar mensaje: {str(e)}")
            else:
                logger.error("No se pudo descargar la imagen del banner")

    except Exception as e:
        logger.error(f"Error al procesar la información de la tarjeta: {str(e)}")
        logger.exception(e)

async def setup_telegram_client() -> TelegramClient:
    """
    Configura y retorna el cliente de Telegram.
    """
    client = TelegramClient('session_name', API_ID, API_HASH)
    
    @client.on(events.NewMessage(chats=SOURCE_CHANNEL))
    async def message_handler(event):
        logger.info(f"Mensaje recibido de {SOURCE_CHANNEL}")
        card_info = extract_card_info(event.raw_text)
        if card_info:
            await process_card_info(client, card_info, bin_database, extras_db)

    return client

async def main():
    """
    Función principal que inicia el bot.
    """
    try:
        # Inicializar las bases de datos
        global bin_database, extras_db
        bin_database = BinDatabase()
        extras_db = ExtrasDatabase()
        
        # Configurar y iniciar el cliente
        client = await setup_telegram_client()
        
        # Ejecutar el cliente
        async with client:
            logger.info("Bot iniciado exitosamente")
            await client.run_until_disconnected()
            
    except Exception as e:
        logger.error(f"Error en la función principal: {str(e)}")
        raise
    finally:
        if 'extras_db' in globals():
            extras_db.close()

if __name__ == '__main__':
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logger.info("Bot detenido por el usuario")
    except Exception as e:
        logger.error(f"Error crítico: {str(e)}")